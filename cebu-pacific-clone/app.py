from flask import Flask, request, jsonify
from flask_cors import CORS
import sqlite3
import hashlib
import re
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)  

EMAIL_REGEX = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'

def get_db_connection():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    return conn

# Initialize SQLite DB
def init_db():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            email TEXT PRIMARY KEY,
            first_name TEXT,
            last_name TEXT,
            password_hash TEXT
        )
    ''')
    conn.commit()
    conn.close()

# Initialize Excel Spreadsheet
def init_excel():
    excel_file = "users.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registered Users"
        ws.append(["Email Address", "First Name", "Last Name", "Password Hash"])
        wb.save(excel_file)

@app.route('/signup', methods=['POST'])
def signup():
    data = request.json
    email = data.get('email', '').strip().lower()
    first_name = data.get('first_name', '').strip()
    last_name = data.get('last_name', '').strip()
    password = data.get('password', '')

    if not email or not re.match(EMAIL_REGEX, email):
        return jsonify({"error": "Invalid email address"}), 400
    if not password or len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters"}), 400

    password_hash = hashlib.sha256(password.encode()).hexdigest()

    # 1. Save to SQLite Database
    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO users (email, first_name, last_name, password_hash) VALUES (?, ?, ?, ?)",
            (email, first_name, last_name, password_hash)
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Email is already registered."}), 400
    conn.close()

    # 2. Save to Excel Spreadsheet
    try:
        wb = load_workbook("users.xlsx")
        ws = wb.active
        ws.append([email, first_name, last_name, password_hash])
        wb.save("users.xlsx")
    except Exception as e:
        print(f"Excel Error: {e}")
        # We still return success because the DB save worked, but we log the Excel error
        return jsonify({"message": "Registration successful (DB only, Excel failed)."}), 201

    return jsonify({"message": "Registration successful! Saved to SQLite and Excel."}), 201

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    password_hash = hashlib.sha256(password.encode()).hexdigest()

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
    conn.close()

    if user is None or user['password_hash'] != password_hash:
        return jsonify({"error": "Invalid email or password."}), 401

    return jsonify({"message": f"Welcome back, {user['first_name']}!"}), 200

if __name__ == '__main__':
    init_db()
    init_excel() # Run the Excel setup when the server starts
    app.run(debug=True, port=5000)